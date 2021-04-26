import requests
from bs4 import BeautifulSoup
import time
import openpyxl
#########################################
year_start = int(input("Enter The Start Year:: "))
year_end = int(input("Enter The End Year:: "))
year_end = year_end + 1
number_of_diary = int(input("Enter The Number Of Diary:: "))
number_of_diary = number_of_diary + 1
print("NOTE:: YOUR DATA WILL BE SAVED TO collected_data.xlsx")
##########################################
loc = 'collected_data.xlsx'
wb = openpyxl.load_workbook(filename=loc)
ws = wb.worksheets[0]
#########################################
for m in range(year_start,year_end):
    try:
        sheet_year_title = "y"+str(m)
        ws_year = wb.create_sheet(sheet_year_title)
        ws_year = wb[sheet_year_title]
        new_m = m-year_start

        for k in range(1,number_of_diary):
            #print(new_m,k)
            computed_row = (3*new_m) + k
            computed_row = abs(computed_row)
            print(computed_row)
            
            response = requests.get("https://main.sci.gov.in/php/captcha_num.php")
            captcha_number = response.json()
            data = {
                "d_no": k,
                "d_yr": m,
                "ansCaptcha": captcha_number,
            }
            headers = {
                #'Host': 'main.sci.gov.in',
                #'Origin': 'https://main.sci.gov.in',
                "Referer": "https://main.sci.gov.in/case-status",
            }
            # Only Referer header is required to bypass CORS
            # You can verify using response.url
            response = requests.post("https://main.sci.gov.in/php/case_status/case_status_process.php", data=data, headers=headers)

            soup=BeautifulSoup(response.text,"html.parser")
            table=soup.find("table")

            try:
                for i in range(0,len(table.contents)):
                    row = BeautifulSoup(str(table.contents[i]),"html.parser")
                    column = row.find_all("td")
                    for j in range(0,len(column)):
                        #print(column[j].text)
                        my_cell_value = column[j].text
                        if j==0:
                            #ws.cell(row=i+1, column=1, value=my_cell_value)
                            pass
                        else:                        
                            try:
                                ws.cell(row=computed_row, column=i+1, value=my_cell_value)
                                ws_year.cell(row=k+1, column=i+1, value=my_cell_value)
                            except Exception as e:
                                print(e)
            except Exception as e:
                print(e,m,k,"Probably Record Does Not Exist")
                print("##############################")
            wb.save(loc)
            time.sleep(3)
            print("##DONE##  ",m,k,i,j)
    except Exception as e:
        print(e,m)
##################################################        










'''
with requests.Session() as session:
    response = session.get("https://main.sci.gov.in/php/captcha_num.php")
    captcha_number = response.json()
    data = {
        "d_no": 1,
        "d_yr": 2000,
        "ansCaptcha": captcha_number,
    }
    headers = {
        #'Host': 'main.sci.gov.in',
        #'Origin': 'https://main.sci.gov.in',
        "Referer": "https://main.sci.gov.in/case-status",
    }
    # Only Referer header is required to bypass CORS
    # You can verify using response.url
    response = session.post("https://main.sci.gov.in/php/case_status/case_status_process.php", data=data, headers=headers)
'''
